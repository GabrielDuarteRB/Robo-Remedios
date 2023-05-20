import requests
from bs4 import BeautifulSoup
import openpyxl
from busca.cliqueFarma import cliqueFarma 
from busca.venancio import venancio
from busca.pagueMenos import pagueMenos

def buscaFiltros(remediosAProcurar, remedios):
    filtros = []
    for row in remediosAProcurar.iter_rows(min_row=2):
        
        remedios[row[2].value]["link"].append(row[1].value)
        filtros = []
        i = 3
        while i < 13 and row[i].value != None:
            filtros.append(row[i].value)       
            i += 1

        remedios[row[2].value]["filtros"].append(filtros)

def iniciar():
    arquivo_excel = openpyxl.Workbook()
    planilha = arquivo_excel.active

    planilhaFiltros = openpyxl.load_workbook('filtros_remedios.xlsx')
    remediosAProcurar = planilhaFiltros['remedios']

    remedios = {
        "pague menos": {"link": [], "filtros": []},
        "clique farma": {"link": [], "filtros": []},
        "venancio": {"link": [], "filtros": []}
    }     

    buscaFiltros(remediosAProcurar, remedios)

    planilha["A1"] = "Nome"
    planilha["B1"] = "Loja"
    planilha["C1"] = "Fabricante"
    planilha["D1"] = "Preço"
    planilha["E1"] = "Link"


    cliqueFarma(planilha, remedios["clique farma"]["link"], remedios["clique farma"]["filtros"])
    venancio(planilha, remedios["venancio"]["link"], remedios["venancio"]["filtros"])
    pagueMenos(planilha, remedios["pague menos"]["link"], remedios["pague menos"]["filtros"])
          
    arquivo_excel.save("remedios.xlsx")
    planilhaFiltros.close()

iniciar()